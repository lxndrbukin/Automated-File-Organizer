import openpyxl
from pathlib import Path

def initialize_log(log_path):
    if not Path.exists(log_path):
        wb = openpyxl.Workbook()
        ws = wb["Sheet"]
        ws.title = "Logs"
        header_row = ws["A1:C1"]
        log_headers = ["File", "Renamed To", "Destination"]
        for i, cell in enumerate(header_row[0]):
            cell.value = log_headers[i]
        wb.save("logs.xlsx")

def log_to_doc(doc, rows):
    log_wb = openpyxl.load_workbook(doc)
    log_ws = log_wb["Logs"]

    for row in rows:
        log_ws.append(row)
    log_wb.save("logs.xlsx")