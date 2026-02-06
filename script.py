from utils import home_dir, config_path, log_path, config, default_src_dir
from datetime import datetime, date
from pathlib import Path
from tabulate import tabulate
import shutil
import json
import openpyxl

default_config = config()

if not Path.exists(config_path):
    action = ""
    while action.lower() not in ["y", "n"]:
        action = input(f"Would you like to use a default 'config.json' file? y/n: ")
        if action.lower() not in ["y", "n"]:
            print("Please enter 'y' or 'n'")
    if action.lower() == "y":
        with open("config.json", "w") as config_file:
            json.dump(default_config, config_file, indent=4)
            print("Default 'config.json' generated.")
    if action.lower() == "n":
        while True:
            src_path = Path(input("Please enter the full source directory for sorting:\n"))
            if not Path.exists(src_path):
                print("The provided source directory doesn't exist. Please enter an existing directory.")
            else:
                break
        target_path = home_dir
        target_dirs_list = input("Please enter target directories, separated by commas (e.g. Images,Videos,Documents):\n")
        confirmation = ""
        while confirmation.lower() not in ["y", "n"]:
            confirmation = input(
                f"Entered directories will be created under {home_dir}.\n Would you like to proceed? y/n: ")
            if confirmation.lower() not in ["y", "n"]:
                print("Please enter 'y' or 'n'")
            if confirmation.lower() == "n":
                while True:
                    target_path = Path(input("Please enter the desired full target path:\n"))
                    if not Path.exists(target_path):
                        print("Provided target path doesn't exist. Please enter an existing path.")
                    else:
                        break
            target_dirs = []
            for category in target_dirs_list.replace(" ", "").split(","):
                formats = input(f"Please enter desired file extensions to be saved in {target_path / category} directory (e.g. xlsx,doc,gif):\n").replace(" ", "").split(",")
                sub_dir_example = Path(target_path / category / f"{category.lower()}_{date.today()}")
                sub_dirs = True
                while True:
                    sub_dir_check = input(
                        f"Would you like the files to be sorted into sub directories based on dates? (e.g. {sub_dir_example}) y/n: ")
                    if sub_dir_check.lower() not in ["y", "n"]:
                        print("Please enter 'y' or 'n'.")
                    else:
                        break
                if sub_dir_check.lower() == "n":
                    sub_dirs = False
                target_dirs.append({f"{category}": {"formats": formats, "sub_dirs": sub_dirs}})
            with open("config.json", "w") as config_file:
                json.dump(config(src_dir=src_path, target_path=target_path, target_dirs=target_dirs), config_file, indent=4)
                print("'config.json' now configured")

with open("config.json", "r") as config_file:
    config = json.load(config_file)

if not Path.exists(Path(config["src_dir"])):
    print(f"Directory {config["src_dir"]} does not exist")
    option = ""
    while option not in ["1", "2"]:
        option = input("Would you like to:\n1. Create the directory\n2. Update src_dir\nPlease enter option number: ")
        if option not in ["1", "2"]:
            print("Please enter 1 or 2")
    if option == "1":
        new_dir = Path(config["src_dir"])
        new_dir.mkdir()
        print("Directory created")
    elif option == "2":
        new_dir = Path(input("Please enter the new directory path: "))
        with open("config.json", "w") as config_file:
            json.dump(config(new_dir), config_file, indent=4)
            print("'config.json' updated!")

if not Path.exists(log_path):
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws.title = "Logs"
    header_row = ws["A1:E1"]
    log_headers = ["Source Dir", "Target Dir", "Old name", "New name", "Date sorted"]
    for i, cell in enumerate(header_row[0]):
        cell.value = log_headers[i]
    wb.save("logs.xlsx")

src_dir = Path(config["src_dir"])

log_rows = []

def log_to_doc(doc, rows):
    log_wb = openpyxl.load_workbook(doc)
    log_ws = log_wb["Logs"]

    for row in rows:
        log_ws.append(row)
    log_wb.save("logs.xlsx")

def sort_to_dir(src_dir, sort_dir, formats, use_sub_dirs):
    sort_dir_path = home_dir / sort_dir
    sort_dir_path.mkdir(exist_ok=True)
    for file in src_dir.iterdir():
        file_date = datetime.fromtimestamp(file.stat().st_mtime).date()
        file_path = src_dir / file.name
        if use_sub_dirs:
            target_dir_path = sort_dir_path / f"{sort_dir.lower()}_{str(file_date)}"
        else:
            target_dir_path = sort_dir_path
        file_name = file.name
        if file.is_file():
            if file.suffix.lower() in formats:
                counter = 0
                target_dir_path.mkdir(exist_ok=True)
                while Path.exists(target_dir_path / file_name):
                    counter += 1
                    file_name = file.stem + f"_{str(counter)}" + file.suffix
                log_row = [str(src_dir), str(target_dir_path), file.name, file_name if file_name != file.name else "", str(file_date)]
                log_rows.append(log_row)
                try:
                    shutil.move(file_path, target_dir_path / file_name)
                except PermissionError as e:
                    print(e)

def run_script():
    try:
        print(f"Sorting '{default_src_dir}'...\n")
        for category_config in config["target_dirs_config"]["target_dirs"]:
            for name, formats in category_config.items():
                formats_list = ["." + fmt if not fmt.startswith(".") else fmt for fmt in formats["formats"]]
                sort_to_dir(src_dir, name, formats_list, formats["sub_dirs"])
        if len(log_rows):
            log_to_doc("logs.xlsx", log_rows)
            print(f"Sorting complete! {len(log_rows)} file(s) organized.\n")
            display_rows = [[row[2], row[3] or "-", Path(row[1]).parts[-1]]
                            for row in log_rows]
            print(tabulate(display_rows, headers=["File", "Renamed To", "Destination Folder"]))
        else:
            print("No files to sort")
    except FileNotFoundError as e:
        print(e)

if __name__ == "__main__":
    run_script()